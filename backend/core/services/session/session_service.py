import json
import os
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
import pandas as pd
import uuid

from django.conf import settings
from django.core.cache import cache

from core.services.error_handler import safe_operation, ApiErrorResponse

logger = logging.getLogger(__name__)

class SessionService:
    """
    Handles user sessions, analysis history, and state management within the application.
    
    This service provides methods for:
    - Tracking analysis history
    - Saving and loading analysis results
    - Managing visualizations
    - Retrieving user history
    
    The service uses a combination of database storage and file system storage
    for efficiency and persistence.
    """
    
    def __init__(self):
        """Initialize session service with necessary storage paths."""
        self._ensure_storage_directories()
        
    def _ensure_storage_directories(self):
        """Ensure required storage directories exist."""
        base_dir = settings.BASE_DIR
        directories = [
            os.path.join(base_dir, "data", "user_sessions"),
            os.path.join(base_dir, "data", "analysis_history"),
            os.path.join(base_dir, "data", "plots"),
            os.path.join(base_dir, "data", "reports")
        ]
        
        for directory in directories:
            Path(directory).mkdir(parents=True, exist_ok=True)
            
    @safe_operation
    def save_analysis_result(self, username: str, analysis_type: str, 
                           data: pd.DataFrame, parameters: Dict, 
                           results: Dict, plots: Optional[List[Dict]] = None) -> bool:
        """
        Save analysis results with associated plots and metadata.
        
        Args:
            username: User identifier
            analysis_type: Type of analysis performed
            data: DataFrame used in the analysis
            parameters: Parameters used for the analysis
            results: Results of the analysis
            plots: List of plot data dictionaries
            
        Returns:
            Boolean indicating success or failure
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        analysis_id = f"{analysis_type}_{timestamp}"
        history_dir = Path(os.path.join(settings.BASE_DIR, "data", "analysis_history", username))
        history_dir.mkdir(parents=True, exist_ok=True)
        history_file = history_dir / "history.json"
        
        # Create base analysis record
        analysis_record = {
            'id': analysis_id,
            'timestamp': timestamp,
            'analysis_type': analysis_type,
            'parameters': parameters,
            'results': self._prepare_results_for_storage(results),
            'data_shape': data.shape if data is not None else None,
            'date_created': datetime.now().isoformat()
        }

        # Handle plot saving
        if plots:
            plots_dir = Path(os.path.join(settings.BASE_DIR, "data", "plots", username, timestamp))
            plots_dir.mkdir(parents=True, exist_ok=True)
            plot_references = []
            
            for idx, plot_data in enumerate(plots):
                plot_path = plots_dir / f"plot_{idx}.json"
                # Save plot data atomically
                with open(plot_path.with_suffix('.tmp'), 'w') as f:
                    json.dump(plot_data.get('plot_data', {}), f)
                os.replace(plot_path.with_suffix('.tmp'), plot_path)
                
                plot_references.append({
                    'id': str(uuid.uuid4()),
                    'path': str(plot_path),
                    'title': plot_data.get('title', ''),
                    'description': plot_data.get('description', ''),
                    'type': plot_data.get('type', 'plot')
                })
            
            analysis_record['plots'] = plot_references

        # Load existing history
        history = []
        if history_file.exists():
            try:
                with open(history_file, 'r') as f:
                    history = json.load(f)
            except json.JSONDecodeError:
                logger.warning(f"Corrupt history file found for {username}, starting fresh")

        # Add new record
        history.append(analysis_record)

        # Save updated history atomically
        with open(history_file.with_suffix('.tmp'), 'w') as f:
            json.dump(history, f, indent=2)
        os.replace(history_file.with_suffix('.tmp'), history_file)
        
        # Store reference in cache for quick access
        cache_key = f"analysis_history_{username}_{analysis_id}"
        cache.set(cache_key, analysis_record, timeout=86400)  # 24 hour cache
        
        return True
    
    def _prepare_results_for_storage(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare results for storage by handling non-serializable objects."""
        clean_results = {}
        
        for key, value in results.items():
            # Skip non-serializable components like model objects
            if key in ['model', 'trace', 'classifier', 'regressor']:
                continue
                
            # Handle nested dictionaries
            if isinstance(value, dict):
                clean_results[key] = self._prepare_results_for_storage(value)
            # Handle numpy arrays and pandas objects
            elif hasattr(value, 'tolist'):
                clean_results[key] = value.tolist()
            elif hasattr(value, 'to_dict'):
                clean_results[key] = value.to_dict()
            # Include serializable values
            elif isinstance(value, (str, int, float, bool, list)) or value is None:
                clean_results[key] = value
            else:
                # Attempt to convert to string for other objects
                try:
                    clean_results[key] = str(value)
                except Exception:
                    logger.warning(f"Skipping non-serializable value for key: {key}")
        
        return clean_results
        
    @safe_operation
    def get_user_history(self, username: str, load_plots: bool = True) -> List[Dict[str, Any]]:
        """
        Get user's analysis history.
        
        Args:
            username: User identifier
            load_plots: Whether to load associated plot data
            
        Returns:
            List of analysis history records
        """
        history_file = Path(os.path.join(settings.BASE_DIR, "data", "analysis_history", username, "history.json"))
        if not history_file.exists():
            return []
            
        with open(history_file, 'r') as f:
            history = json.load(f)
            
        if load_plots:
            return self._load_plots_for_history(username, history)
        
        return history
    
    @safe_operation
    def get_analysis_by_id(self, username: str, analysis_id: str) -> Optional[Dict[str, Any]]:
        """
        Get a specific analysis by ID.
        
        Args:
            username: User identifier
            analysis_id: Analysis identifier
            
        Returns:
            Analysis record if found, None otherwise
        """
        # Check cache first
        cache_key = f"analysis_history_{username}_{analysis_id}"
        cached_result = cache.get(cache_key)
        if cached_result:
            return cached_result
        
        # If not in cache, look in file system
        history = self.get_user_history(username, load_plots=False)
        for record in history:
            if record.get('id') == analysis_id:
                # Store in cache for future requests
                cache.set(cache_key, record, timeout=86400)  # 24 hour cache
                return record
                
        return None
    
    @safe_operation
    def _load_plots_for_history(self, username: str, history: List[Dict]) -> List[Dict]:
        """
        Load plot data for history records.
        
        Args:
            username: User identifier
            history: List of history records
            
        Returns:
            History records with loaded plot data
        """
        for record in history:
            if 'plots' in record and isinstance(record['plots'], list):
                for i, plot in enumerate(record['plots']):
                    if 'path' in plot and os.path.exists(plot['path']):
                        try:
                            with open(plot['path'], 'r') as f:
                                plot_data = json.load(f)
                            record['plots'][i]['plot_data'] = plot_data
                        except Exception as e:
                            logger.error(f"Error loading plot data: {str(e)}")
        return history
    
    @safe_operation
    def clear_user_history(self, username: str) -> bool:
        """
        Clear user's analysis history.
        
        Args:
            username: User identifier
            
        Returns:
            Boolean indicating success or failure
        """
        # Delete history file
        history_file = Path(os.path.join(settings.BASE_DIR, "data", "analysis_history", username, "history.json"))
        if history_file.exists():
            os.remove(history_file)
        
        # Delete plot files
        plots_dir = Path(os.path.join(settings.BASE_DIR, "data", "plots", username))
        if plots_dir.exists():
            import shutil
            shutil.rmtree(plots_dir)
        
        # Clear cache for this user
        cache.delete_pattern(f"analysis_history_{username}_*")
        
        return True
    
    @safe_operation    
    def get_analysis_statistics(self, username: str) -> Dict[str, Any]:
        """
        Get statistics about user's analysis history.
        
        Args:
            username: User identifier
            
        Returns:
            Dictionary of statistics
        """
        history = self.get_user_history(username, load_plots=False)
        
        if not history:
            return {
                'total_analyses': 0,
                'unique_analysis_types': [],
                'total_plots': 0,
                'latest_analysis': None,
                'analysis_counts': {}
            }
            
        # Calculate statistics
        analysis_types = [record['analysis_type'] for record in history]
        unique_types = list(set(analysis_types))
        type_counts = {t: analysis_types.count(t) for t in unique_types}
        
        total_plots = sum(len(record.get('plots', [])) for record in history)
        
        # Get latest analysis
        latest_analysis = max(history, key=lambda x: x.get('timestamp', ''))
        
        return {
            'total_analyses': len(history),
            'unique_analysis_types': unique_types,
            'total_plots': total_plots,
            'latest_analysis': {
                'id': latest_analysis.get('id'),
                'type': latest_analysis.get('analysis_type'),
                'timestamp': latest_analysis.get('timestamp')
            },
            'analysis_counts': type_counts
        }
        
    @safe_operation
    def export_analysis_data(self, username: str, analysis_id: str, format: str = 'json') -> Tuple[Optional[bytes], str]:
        """
        Export analysis data in specified format.
        
        Args:
            username: User identifier
            analysis_id: Analysis identifier
            format: Export format ('json', 'csv', 'xlsx')
            
        Returns:
            Tuple of (data bytes, filename)
        """
        analysis = self.get_analysis_by_id(username, analysis_id)
        if not analysis:
            return None, ''
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == 'json':
            data = json.dumps(analysis, indent=2).encode('utf-8')
            filename = f"analysis_{analysis_id}_{timestamp}.json"
            
        elif format == 'csv':
            # Create a flattened CSV representation
            flattened_data = self._flatten_analysis_data(analysis)
            
            import io
            csv_buffer = io.StringIO()
            pd.DataFrame([flattened_data]).to_csv(csv_buffer, index=False)
            
            data = csv_buffer.getvalue().encode('utf-8')
            filename = f"analysis_{analysis_id}_{timestamp}.csv"
            
        elif format == 'xlsx':
            # Create an Excel workbook with multiple sheets
            import io
            from openpyxl import Workbook
            
            wb = Workbook()
            
            # Create main sheet
            ws = wb.active
            ws.title = "Analysis Overview"
            
            ws['A1'] = "Analysis ID"
            ws['B1'] = analysis_id
            ws['A2'] = "Analysis Type"
            ws['B2'] = analysis.get('analysis_type', '')
            ws['A3'] = "Timestamp"
            ws['B3'] = analysis.get('timestamp', '')
            
            # Create parameters sheet
            if 'parameters' in analysis:
                param_sheet = wb.create_sheet("Parameters")
                for i, (key, value) in enumerate(analysis['parameters'].items(), 1):
                    param_sheet[f'A{i}'] = key
                    param_sheet[f'B{i}'] = str(value)
            
            # Create results sheet
            if 'results' in analysis:
                results_sheet = wb.create_sheet("Results")
                flattened_results = self._flatten_analysis_data(analysis['results'])
                for i, (key, value) in enumerate(flattened_results.items(), 1):
                    results_sheet[f'A{i}'] = key
                    results_sheet[f'B{i}'] = str(value)
            
            # Save to buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            data = excel_buffer.read()
            filename = f"analysis_{analysis_id}_{timestamp}.xlsx"
            
        else:
            raise ValueError(f"Unsupported format: {format}")
            
        return data, filename
    
    def _flatten_analysis_data(self, data: Dict[str, Any], prefix: str = '') -> Dict[str, Any]:
        """
        Flatten nested dictionaries for export formats like CSV.
        
        Args:
            data: Dictionary to flatten
            prefix: Prefix for nested keys
            
        Returns:
            Flattened dictionary
        """
        flattened = {}
        
        for key, value in data.items():
            new_key = f"{prefix}.{key}" if prefix else key
            
            if isinstance(value, dict):
                nested_flattened = self._flatten_analysis_data(value, new_key)
                flattened.update(nested_flattened)
            elif isinstance(value, list):
                # For lists, store length and first few items
                flattened[new_key + '.length'] = len(value)
                for i, item in enumerate(value[:3]):
                    flattened[new_key + f'.item{i}'] = str(item)
                if len(value) > 3:
                    flattened[new_key + '.truncated'] = True
            else:
                flattened[new_key] = value
                
        return flattened

# Initialize global session service
session_service = SessionService()

def get_session_service() -> SessionService:
    """Get the global session service instance."""
    return session_service